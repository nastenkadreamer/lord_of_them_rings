#!/usr/bin/env python3
"""
MP Village Shapefile - LGD Crosswalk Matching Pipeline (v4)
============================================================
Matches Survey of India village shapefiles to LGD crosswalk records
for all Madhya Pradesh districts.

v4 changes:
    - District name aliases (HOSHANGABAD→Narmadapuram, WEST_NIMAR→Khargone,
      EAST_NIMAR→Khandwa, NARSHIMAPURA→Narsimhapur, NIMACH→Neemuch)
    - Parent→child district merging (Satna+Maihar, Rewa+Mauganj,
      Tikamgarh+Niwari, Chhindwara+Pandhurna)
    - Land tenure suffix stripping (Ryt., Mal., Kh., Bk.) — fixes Dindori
      (80% of names have these suffixes), Mandla (25%), Khandwa (17%)

v3 changes (two-pass inference):
    - Pass 1: exact-match district-wide to learn tehsil→blocks mapping
    - Pass 2: constrained fuzzy using inferred blocks per tehsil
    - Added ! → U to SoI font map, PH↔F substitution
    - Improved RF/PF detection
"""

import os, sys, re, math, struct, argparse, logging
from pathlib import Path
from difflib import SequenceMatcher

try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas required"); sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

CHAR_MAP = {'>':'A', '|':'I', '@':'U', '<':'E', '!':'U'}
TRANS_RULES = [
    (r'KHERI\b','KHEDI'), (r'KHERA\b','KHEDA'),
    (r'BARARIYA\b','BARDIYA'),
    (r'PIPLYA\b','PIPALIYA'), (r'PIPLIYA\b','PIPALIYA'),
    (r'CHOPRA\b','CHOPDA'), (r'CHAPRA\b','CHOPDA'),
]
AUTO_THRESH = 0.85
LCC = dict(a=6378137.0, f=1/298.257223563, lon0=80.0, lat0=24.0,
           lat1=12.472944, lat2=35.172806, FE=4e6, FN=4e6)

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s', datefmt='%H:%M:%S')
log = logging.getLogger('village_match')

# --- Projection ---
class LCCProj:
    def __init__(self, p=None):
        p=p or LCC; self.a=p['a']; self.e=math.sqrt(2*p['f']-p['f']**2)
        self.lon0=math.radians(p['lon0']); self.FE=p['FE']; self.FN=p['FN']
        l0,l1,l2=[math.radians(p[k]) for k in ('lat0','lat1','lat2')]
        m1,m2=self._m(l1),self._m(l2); t0,t1,t2=self._t(l0),self._t(l1),self._t(l2)
        self.n=(math.log(m1)-math.log(m2))/(math.log(t1)-math.log(t2))
        self.F=m1/(self.n*t1**self.n); self.rho0=self.a*self.F*t0**self.n
    def _m(self,lat): return math.cos(lat)/math.sqrt(1-self.e**2*math.sin(lat)**2)
    def _t(self,lat):
        es=self.e*math.sin(lat); return math.tan(math.pi/4-lat/2)/((1-es)/(1+es))**(self.e/2)
    def to_ll(self,ex,ny):
        x,y=ex-self.FE,ny-self.FN; rho=math.sqrt(x**2+(self.rho0-y)**2)
        t=(rho/(self.a*self.F))**(1/self.n); th=math.atan2(x,self.rho0-y)
        lon=th/self.n+self.lon0; lat=math.pi/2-2*math.atan(t)
        for _ in range(10):
            es=self.e*math.sin(lat)
            lat2=math.pi/2-2*math.atan(t*((1-es)/(1+es))**(self.e/2))
            if abs(lat2-lat)<1e-12: break
            lat=lat2
        return round(math.degrees(lon),6), round(math.degrees(lat),6)

# --- Shapefile parsing ---
def read_dbf(path):
    rows=[]
    with open(path,'rb') as f:
        f.read(4); nr=struct.unpack('<I',f.read(4))[0]
        hs=struct.unpack('<H',f.read(2))[0]; f.read(22)
        fields=[]
        for _ in range((hs-33)//32):
            fd=f.read(32)
            if fd[0]==0x0D: break
            fields.append((fd[:11].replace(b'\x00',b'').decode('ascii'),fd[16]))
        f.seek(hs)
        for _ in range(nr):
            f.read(1); row={}
            for name,fsize in fields:
                raw=f.read(fsize)
                try: val=raw.decode('utf-8').strip()
                except:
                    try: val=raw.decode('latin-1').strip()
                    except: val=''
                row[name]=val
            rows.append(row)
    return rows

def read_centroids(path):
    cents=[]
    with open(path,'rb') as f:
        f.read(24); fl=struct.unpack('>I',f.read(4))[0]*2; f.read(72)
        while f.tell()<fl:
            try:
                f.read(4); cl=struct.unpack('>I',f.read(4))[0]*2; st=f.tell()
                stype=struct.unpack('<I',f.read(4))[0]
                if stype==5:
                    xmin,ymin,xmax,ymax=struct.unpack('<4d',f.read(32))
                    cents.append(((xmin+xmax)/2,(ymin+ymax)/2))
                else: cents.append((None,None))
                f.seek(st+cl)
            except: cents.append((None,None)); break
    return cents

# --- Text normalization ---
def decode_font(n):
    s=n.upper().strip()
    for o,v in CHAR_MAP.items(): s=s.replace(o,v)
    return re.sub(r'\s+',' ',s).strip()

def norm_trans(n):
    s=n
    for p,r in TRANS_RULES: s=re.sub(p,r,s)
    return s

def strip_suf(n, suf):
    parts=n.split()
    return ' '.join(parts[:-1]) if len(parts)>1 and parts[-1] in suf else n

def strip_rfpf(n):
    return re.sub(r'\s*(R\s*F|P\s*F|PROTECTED\s*FOREST)\s*$','',n).strip()

def strip_tenure(n):
    """Strip land tenure suffixes: Ryt., Mal., Kh., Bk. (common in central MP districts)."""
    return re.sub(r'\s+(RYT\.?|MAL\.?|KH\.?|BK\.?)\s*$', '', n, flags=re.IGNORECASE).strip()

def is_rfpf(n):
    d=decode_font(n).strip().upper()
    return bool(re.search(r'\bR\.?\s*F\.?\b|\bP\.?\s*F\.?\b|PROTECTED\s*FOREST|RESERVED\s*FOREST', d))

def sim(a,b): return SequenceMatcher(None,a,b).ratio()

def sim_rd(a,b):
    """Enhanced similarity with R<->D, V<->W, OU<->O, PH<->F substitutions."""
    scores=[sim(a,b)]
    for s,d in [('R','D'),('D','R'),('V','W'),('W','V'),('PH','F'),('F','PH')]:
        scores.append(sim(re.sub(s,d,a),b)); scores.append(sim(a,re.sub(s,d,b)))
    scores.append(sim(a.replace('OU','O'),b)); scores.append(sim(a,b.replace('OU','O')))
    return max(scores)

def multi_sim(shp_name, cw_row, suf):
    """Best similarity across all name variants + R<->D."""
    cvs=[cw_row['name_upper'],cw_row['name_norm'],cw_row['name_str'],cw_row['name_str_norm'],
         cw_row.get('name_tenure',cw_row['name_upper']),cw_row.get('name_tenure_norm',cw_row['name_norm'])]
    svs=[shp_name, norm_trans(shp_name), strip_suf(shp_name,suf),
         norm_trans(strip_suf(shp_name,suf))]
    best=0
    for sv in svs:
        for cv in cvs:
            s=sim_rd(sv,cv)
            if s>best: best=s
    return best

# --- Tehsil->Block mapping ---
def build_tb_map(shp_rows, cw_dist):
    tehsils=set()
    for row in shp_rows:
        t=row.get('TEHSIL','').strip()
        if t: tehsils.add(t.upper())
    blocks=set(cw_dist['LG_block_name'].str.upper().str.strip().unique())
    mapping={}; used=set()
    for t in sorted(tehsils):
        for b in blocks:
            if t==b: mapping[t]=b; used.add(b); break
    for t in sorted(tehsils-set(mapping.keys())):
        best_s=0; best_b=None
        for b in blocks-used:
            s=sim_rd(t,b)
            if s>best_s: best_s=s; best_b=b
        if best_b and best_s>=0.50:
            mapping[t]=best_b; used.add(best_b)
        else: log.warning(f"  Tehsil '{t}' unmapped (best={best_s:.2f})")
    return mapping

# --- Town detection ---
def detect_towns(shp_rows, cw_dist, tb_map):
    admin=set(tb_map.keys())|set(tb_map.values())
    for b in cw_dist['LG_block_name'].str.upper().str.strip().unique(): admin.add(b)
    cw_names=set(cw_dist['LG_vill_name'].str.upper().str.strip())
    towns=set()
    for idx,row in enumerate(shp_rows):
        base=strip_rfpf(decode_font(row.get('VILLAGE','')))
        if base and base in admin and base not in cw_names: towns.add(idx)
    return towns

# --- Core matching ---
def match_district(dist_name, shp_rows, centroids, cw_dist):
    log.info(f"  {dist_name}: {len(shp_rows)} polygons, {len(cw_dist)} cw villages")
    cw=cw_dist.copy()
    cw['name_upper']=cw['LG_vill_name'].str.upper().str.strip()
    cw['block_upper']=cw['LG_block_name'].str.upper().str.strip()
    suf=set(cw['block_upper'].unique())|{'KANAD','KALAN','KHURD','BUZURG','KHEDA','KHEDI','BADOD','KHAS','BUZRUK','KURD','BUJURG'}
    cw['name_norm']=cw['name_upper'].apply(norm_trans)
    cw['name_str']=cw['name_upper'].apply(lambda x: strip_suf(x,suf))
    cw['name_str_norm']=cw['name_str'].apply(norm_trans)
    # Strip land tenure suffixes (Ryt., Mal., Kh., Bk.)
    cw['name_tenure']=cw['name_upper'].apply(strip_tenure)
    cw['name_tenure_norm']=cw['name_tenure'].apply(norm_trans)

    tb_map=build_tb_map(shp_rows,cw_dist)
    log.info(f"  Tehsil map (fuzzy): {tb_map}")
    towns=detect_towns(shp_rows,cw_dist,tb_map)
    log.info(f"  Towns: {len(towns)}")
    proj=LCCProj()

    # =========================================================================
    # PASS 1 (DISCOVERY): Exact match district-wide to learn tehsil→blocks
    # =========================================================================
    from collections import defaultdict as _ddict
    discovery={}; disc_reserved=set()
    tehsil_block_votes=_ddict(lambda: _ddict(int))  # tehsil → {block: count}

    for idx,row in enumerate(shp_rows):
        if idx in towns: continue
        base=strip_rfpf(decode_font(row.get('VILLAGE','')))
        if not base: continue
        tehsil=row.get('TEHSIL','').upper().strip()
        bn=norm_trans(base); bs=strip_suf(base,suf); bsn=norm_trans(bs)
        brd=re.sub(r'R','D',bn); bdr=re.sub(r'D','R',bn)
        # Also PH<->F variants
        bpf=bn.replace('PH','F'); bfp=bn.replace('F','PH')
        variants={base,bn,bs,bsn,brd,bdr,bpf,bfp}
        cands=cw; found=False
        for col in ['name_upper','name_norm','name_str','name_str_norm','name_tenure','name_tenure_norm']:
            for sv in variants:
                ex=cands[cands[col]==sv]
                if len(ex)>=1:
                    m=ex.iloc[0]
                    if m['LG_vill_code'] not in disc_reserved:
                        discovery[idx]={'lg_code':m['LG_vill_code'],'lg_name':m['LG_vill_name'],
                            'gp_code':m['LG_GP_code'],'gp_name':m['LG_GP_name'],
                            'block':m['block_upper'],'match_type':'exact','score':1.0}
                        disc_reserved.add(m['LG_vill_code'])
                        if tehsil:
                            tehsil_block_votes[tehsil][m['block_upper']]+=1
                        found=True; break
            if found: break
    log.info(f"  Discovery (exact, district-wide): {len(discovery)}")

    # Build inferred tehsil → blocks mapping
    # Include any block that got ≥1 vote from a tehsil
    inferred_tb={}
    all_blocks=set(cw['block_upper'].unique())
    for tehsil, votes in tehsil_block_votes.items():
        blocks=set(blk for blk, cnt in votes.items() if cnt>=1)
        if not blocks:
            blocks=all_blocks
        inferred_tb[tehsil]=blocks

    # Log the inferred mapping
    for t in sorted(inferred_tb.keys()):
        blks=inferred_tb[t]
        votes_str=', '.join(f"{b}({tehsil_block_votes[t][b]})" for b in sorted(blks))
        log.info(f"    {t} → [{votes_str}]")

    # For tehsils with NO exact matches, fall back to district-wide
    all_tehsils=set()
    for row in shp_rows:
        t=row.get('TEHSIL','').upper().strip()
        if t: all_tehsils.add(t)
    for t in all_tehsils - set(inferred_tb.keys()):
        inferred_tb[t]=all_blocks
        log.warning(f"    {t} → ALL BLOCKS (no discovery matches)")

    # Now promote discovery matches as our exact matches
    exact=discovery; reserved=disc_reserved

    # =========================================================================
    # PASS 2 (CONSTRAINED FUZZY): Use inferred tehsil→blocks mapping
    # =========================================================================
    # Build per-block first-letter index for fast lookup
    block_fl_index=_ddict(lambda: _ddict(list))  # block → {first_letter → [cw indices]}
    for cidx,c in cw.iterrows():
        nm=c['name_upper']; blk=c['block_upper']
        if nm:
            block_fl_index[blk][nm[0]].append(cidx)
            if nm[0]=='R': block_fl_index[blk]['D'].append(cidx)
            elif nm[0]=='D': block_fl_index[blk]['R'].append(cidx)
            elif nm[0]=='F': block_fl_index[blk]['P'].append(cidx)
            elif nm[0]=='P': block_fl_index[blk]['F'].append(cidx)

    remaining=[i for i in range(len(shp_rows)) if i not in exact and i not in towns]
    fuzz=[]
    for idx in remaining:
        row=shp_rows[idx]
        base=strip_rfpf(decode_font(row.get('VILLAGE','')))
        rf=is_rfpf(row.get('VILLAGE',''))
        tehsil=row.get('TEHSIL','').upper().strip()
        allowed_blocks=inferred_tb.get(tehsil, all_blocks)
        if not base:
            fuzz.append({'idx':idx,'base':base,'block':'','rf':rf,'scores':[],'best':0}); continue
        blen=len(base)
        fl=base[0] if base else ''
        # Gather candidates from allowed blocks only
        cand_idxs=set()
        for blk in allowed_blocks:
            if fl: cand_idxs.update(block_fl_index[blk].get(fl,[]))
            if fl=='R': cand_idxs.update(block_fl_index[blk].get('D',[]))
            elif fl=='D': cand_idxs.update(block_fl_index[blk].get('R',[]))
            elif fl=='P': cand_idxs.update(block_fl_index[blk].get('F',[]))
            elif fl=='F': cand_idxs.update(block_fl_index[blk].get('P',[]))
        scores=[]
        for cidx in cand_idxs:
            c=cw.loc[cidx]
            if abs(len(c['name_upper'])-blen)>5: continue
            s=multi_sim(base,c,suf)
            if s>=0.60:
                scores.append((s,c['LG_vill_name'],c['LG_vill_code'],c['LG_GP_code'],c['LG_GP_name'],c['block_upper']))
        scores.sort(key=lambda x:-x[0])
        best_block=scores[0][5] if scores else ''
        fuzz.append({'idx':idx,'base':base,'block':best_block,'rf':rf,'scores':scores,
                     'best':scores[0][0] if scores else 0})
    fuzz.sort(key=lambda x:-x['best'])

    fuzzy_auto={}; needs_review=[]
    for fc in fuzz:
        idx=fc['idx']
        if not fc['scores']: needs_review.append(fc); continue
        top3=[(s,n,c,g,gn,bl) for s,n,c,g,gn,bl in fc['scores'] if c not in reserved][:3]
        best=top3[0] if top3 else None
        if best and best[0]>=AUTO_THRESH:
            fuzzy_auto[idx]={'lg_code':best[2],'lg_name':best[1],'gp_code':best[3],
                'gp_name':best[4],'block':best[5],'match_type':'fuzzy_auto','score':round(best[0],4)}
            reserved.add(best[2])
        else: fc['top3']=top3; needs_review.append(fc)
    log.info(f"  Fuzzy auto: {len(fuzzy_auto)}, Review pool: {len(needs_review)}")

    # PASS 3: RF share
    rf_shares={}; still_review=[]
    for fc in needs_review:
        idx=fc['idx']
        if fc.get('rf') and fc['base']:
            found=False
            for aidx,info in {**exact,**fuzzy_auto}.items():
                abase=strip_rfpf(decode_font(shp_rows[aidx].get('VILLAGE','')))
                if fc['base']==abase and info['block']==fc['block']:
                    rf_shares[idx]={**info,'match_type':'rf_share'}; found=True; break
            if not found: still_review.append(fc)
        else: still_review.append(fc)
    log.info(f"  RF share: {len(rf_shares)}")

    # PASS 4: PF nearest
    pf_idxs=[]; final_review=[]
    for fc in still_review:
        base=fc.get('base','')
        if not base or base in ('PF','PROTECTED FOREST','P F','R F'): pf_idxs.append(fc['idx'])
        else: final_review.append(fc)
    mcents={}
    for src in [exact,fuzzy_auto,rf_shares]:
        for midx,info in src.items():
            if info.get('lg_code') and midx<len(centroids):
                cx,cy=centroids[midx]
                if cx is not None: mcents[midx]=(cx,cy,info)
    pf_merged={}
    for pidx in pf_idxs:
        if pidx>=len(centroids): continue
        px,py=centroids[pidx]
        if px is None: continue
        bd=float('inf'); bi=None
        for midx,(mx,my,mi) in mcents.items():
            d=math.sqrt((px-mx)**2+(py-my)**2)
            if d<bd: bd=d; bi=mi
        if bi: pf_merged[pidx]={**bi,'match_type':'pf_nearest'}
    log.info(f"  PF nearest: {len(pf_merged)}")

    # Compile
    results=[]
    for idx in range(len(shp_rows)):
        row=shp_rows[idx]
        cx,cy=centroids[idx] if idx<len(centroids) else (None,None)
        try: lon,lat=proj.to_ll(cx,cy) if cx else (None,None)
        except: lon,lat=None,None
        info=(exact.get(idx) or fuzzy_auto.get(idx) or rf_shares.get(idx) or pf_merged.get(idx))
        if info is None:
            mt='skip_town' if idx in towns else 'unmatched'
            info={'lg_code':None,'lg_name':'','gp_code':None,'gp_name':'','block':'','match_type':mt,'score':0}
        results.append({
            'shp_idx':idx,'shp_name':row.get('VILLAGE',''),
            'cleaned':strip_rfpf(decode_font(row.get('VILLAGE',''))),
            'tehsil':row.get('TEHSIL',''),'district':row.get('DISTRICT',dist_name),
            'match_type':info['match_type'],
            'LG_vill_code':info.get('lg_code'),'LG_vill_name':info.get('lg_name',''),
            'LG_GP_code':info.get('gp_code'),'LG_GP_name':info.get('gp_name',''),
            'block':info.get('block',''),'score':info.get('score',0),'lat':lat,'lon':lon,
        })

    review=[]
    for fc in final_review:
        idx=fc['idx']; row=shp_rows[idx]
        cx,cy=centroids[idx] if idx<len(centroids) else (None,None)
        try: lon,lat=proj.to_ll(cx,cy) if cx else (None,None)
        except: lon,lat=None,None
        top3=fc.get('top3',[])
        review.append({
            'shp_idx':idx,'shp_name':row.get('VILLAGE',''),
            'cleaned':fc.get('base',''),'tehsil':row.get('TEHSIL',''),'block':fc.get('block',''),
            'cand1_name':top3[0][1] if len(top3)>0 else '','cand1_score':round(top3[0][0],3) if len(top3)>0 else 0,
            'cand1_gp':top3[0][4] if len(top3)>0 else '',
            'cand2_name':top3[1][1] if len(top3)>1 else '','cand2_score':round(top3[1][0],3) if len(top3)>1 else 0,
            'cand3_name':top3[2][1] if len(top3)>2 else '','cand3_score':round(top3[2][0],3) if len(top3)>2 else 0,
            'lat':lat,'lon':lon,'decision':'',
        })

    df_r=pd.DataFrame(results); matched=df_r['LG_vill_code'].notna().sum()
    stats={
        'district':dist_name,'shp_polygons':len(shp_rows),'cw_villages':len(cw_dist),
        'exact':len(exact),'fuzzy_auto':len(fuzzy_auto),'rf_share':len(rf_shares),
        'pf_nearest':len(pf_merged),'skip_town':len(towns),
        'unmatched':len(df_r[df_r['match_type']=='unmatched']),
        'review_items':len(review),'total_matched':int(matched),
        'match_rate':round(matched/len(shp_rows)*100,1) if shp_rows else 0,
        'unique_villages':int(df_r[df_r['LG_vill_code'].notna()]['LG_vill_code'].nunique()),
        'unique_gps':int(df_r[df_r['LG_GP_code'].notna()]['LG_GP_code'].nunique()),
        'tehsil_block_map':str(tb_map),
    }
    return results, review, stats

# --- File discovery ---
def find_shapefiles(d):
    d=Path(d); out={}
    for shp in sorted(d.rglob('*.shp')):
        dbf=shp.with_suffix('.dbf')
        if dbf.exists(): out[shp.stem.upper()]={'shp':str(shp),'dbf':str(dbf)}
    return out

# --- Output ---
def write_outputs(results, review, stats, out_dir, dist_name):
    out=Path(out_dir); out.mkdir(parents=True,exist_ok=True)
    pd.DataFrame(results).to_csv(out/f'{dist_name}_matched.csv',index=False)
    if review: pd.DataFrame(review).to_csv(out/f'{dist_name}_review.csv',index=False)

def write_summary(all_stats, out_dir):
    out=Path(out_dir); df=pd.DataFrame(all_stats)
    cols=['district','shp_polygons','cw_villages','exact','fuzzy_auto','rf_share',
          'pf_nearest','skip_town','unmatched','review_items','total_matched',
          'match_rate','unique_villages','unique_gps']
    df=df[[c for c in cols if c in df.columns]]
    df.to_csv(out/'mp_match_summary.csv',index=False)

    if HAS_OPENPYXL:
        wb=Workbook(); hdr_fill=PatternFill('solid',fgColor='2F5496')
        hdr_font=Font(bold=True,color='FFFFFF',name='Arial',size=10)
        bf=Font(name='Arial',size=10); thin=Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
        ws=wb.active; ws.title='Summary'
        for c,h in enumerate(list(df.columns),1):
            cell=ws.cell(row=1,column=c,value=h); cell.fill=hdr_fill; cell.font=hdr_font; cell.border=thin
        for r,(_,row) in enumerate(df.iterrows(),2):
            for c,h in enumerate(list(df.columns),1):
                cell=ws.cell(row=r,column=c,value=row[h]); cell.font=bf; cell.border=thin
        r=len(df)+2; ws.cell(row=r,column=1,value='TOTAL').font=Font(bold=True,name='Arial',size=10)
        for c,h in enumerate(list(df.columns),1):
            if h not in ('district','match_rate','tehsil_block_map'):
                try: cell=ws.cell(row=r,column=c,value=int(df[h].sum())); cell.font=Font(bold=True,name='Arial',size=10); cell.border=thin
                except: pass
        for i in range(1,len(df.columns)+1): ws.column_dimensions[get_column_letter(i)].width=max(len(str(list(df.columns)[i-1]))+4,12)
        ws2=wb.create_sheet('TehsilBlockMaps')
        for c,h in enumerate(['District','Tehsil','Block'],1):
            ws2.cell(row=1,column=c,value=h).fill=hdr_fill; ws2.cell(row=1,column=c).font=hdr_font
        r=2
        for st in all_stats:
            try: tbm=eval(st.get('tehsil_block_map','{}'))
            except: tbm={}
            for t,b in sorted(tbm.items()):
                ws2.cell(row=r,column=1,value=st['district']).font=bf
                ws2.cell(row=r,column=2,value=t).font=bf
                ws2.cell(row=r,column=3,value=b).font=bf; r+=1
        ws2.column_dimensions['A'].width=25; ws2.column_dimensions['B'].width=20; ws2.column_dimensions['C'].width=20
        wb.save(out/'mp_match_summary.xlsx')

    print("\n"+"="*80); print("  MP VILLAGE MATCHING SUMMARY"); print("="*80)
    print(f"  Districts: {len(df)}"); print(f"  Total polygons:  {df['shp_polygons'].sum()}")
    print(f"  Total matched:   {df['total_matched'].sum()}")
    if df['shp_polygons'].sum()>0:
        print(f"  Match rate:      {df['total_matched'].sum()/df['shp_polygons'].sum()*100:.1f}%")
    print(f"  Review items:    {df['review_items'].sum()}\n")
    print(f"  {'District':<28} {'Poly':>6} {'Match':>6} {'Rate':>6} {'Review':>7}")
    print(f"  {'-'*28} {'-'*6} {'-'*6} {'-'*6} {'-'*7}")
    for _,r in df.sort_values('district').iterrows():
        print(f"  {r['district']:<28} {r['shp_polygons']:>6} {r['total_matched']:>6} {r['match_rate']:>5.1f}% {r['review_items']:>7}")

# --- Main ---
def main():
    parser=argparse.ArgumentParser(description='MP Village Shapefile Matching Pipeline')
    parser.add_argument('--shapefile-dir',required=True)
    parser.add_argument('--crosswalk',required=True)
    parser.add_argument('--output-dir',required=True)
    parser.add_argument('--districts',nargs='*',default=None)
    parser.add_argument('--auto-threshold',type=float,default=0.85)
    args=parser.parse_args()

    global AUTO_THRESH; AUTO_THRESH=args.auto_threshold

    log.info("Loading crosswalk...")
    cw=pd.read_stata(args.crosswalk)
    mp=cw[cw['LG_state_name'].str.contains('Madhya',na=False,case=False)].copy()
    log.info(f"  {len(mp)} MP villages, {mp['LG_district_name'].nunique()} districts")

    log.info("Scanning shapefiles...")
    dist_files=find_shapefiles(args.shapefile_dir)
    log.info(f"  Found {len(dist_files)} districts")

    if args.districts:
        req=set(d.upper().replace(' ','_').replace('-','_') for d in args.districts)
        dist_files={k:v for k,v in dist_files.items() if k in req}
    if not dist_files: log.error("No shapefiles found!"); sys.exit(1)

    cw_dists=mp['LG_district_name'].unique()

    # Hard-coded aliases: shapefile name → crosswalk district name
    # For renamed districts and spelling mismatches
    DISTRICT_ALIASES = {
        'HOSHANGABAD': 'Narmadapuram',
        'WEST_NIMAR': 'Khargone (West Nimar)',
        'EAST_NIMAR': 'Khandwa (East Nimar)',
        'NARSHIMAPURA': 'Narsimhapur',
        'NIMACH': 'Neemuch',
    }

    # Parent → child districts (newer districts carved from older ones)
    # When matching a parent shapefile, also search child crosswalk villages
    PARENT_CHILDREN = {
        'Satna': ['Maihar'],
        'Rewa': ['Mauganj'],
        'Tikamgarh': ['Niwari'],
        'Chhindwara': ['Pandhurna'],
    }

    def match_cw_dist(name):
        # Check alias table first
        if name in DISTRICT_ALIASES:
            return DISTRICT_ALIASES[name]
        n=name.replace('_',' ').replace('-',' ').upper()
        best_s=0; best=None
        for d in cw_dists:
            s=sim_rd(n,d.upper())
            if s>best_s: best_s=s; best=d
        return best if best_s>=0.60 else None

    out=Path(args.output_dir); out.mkdir(parents=True,exist_ok=True); all_stats=[]
    for shp_dist,paths in sorted(dist_files.items()):
        log.info(f"\n{'='*60}\n  {shp_dist}\n{'='*60}")
        cw_name=match_cw_dist(shp_dist)
        if not cw_name: log.warning(f"  No crosswalk match for '{shp_dist}'"); continue
        log.info(f"  -> {cw_name}")
        try: rows=read_dbf(paths['dbf']); cents=read_centroids(paths['shp'])
        except Exception as e: log.error(f"  Read error: {e}"); continue
        # Include child district villages for parent shapefiles
        child_dists = PARENT_CHILDREN.get(cw_name, [])
        cw_sub=mp[mp['LG_district_name'].isin([cw_name] + child_dists)].copy()
        if child_dists:
            log.info(f"  Including child districts: {child_dists} (+{len(cw_sub)-len(mp[mp['LG_district_name']==cw_name])} villages)")
        try: results,review,stats=match_district(cw_name,rows,cents,cw_sub)
        except Exception as e: log.error(f"  Match error: {e}"); import traceback; traceback.print_exc(); continue
        write_outputs(results,review,stats,out/'districts',shp_dist)
        all_stats.append(stats)
        log.info(f"  -> {stats['total_matched']}/{stats['shp_polygons']} ({stats['match_rate']}%), review={stats['review_items']}")

    if all_stats: write_summary(all_stats,out)
    else: log.error("No districts processed!")

if __name__=='__main__': main()
